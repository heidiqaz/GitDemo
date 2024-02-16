import openpyxl

book = openpyxl.load_workbook("/Users/heidi/PycharmProjects/pythonSelFramework/pyDemo.xlsx")
sheet = book.active
dict = {}
cell = sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "heidi"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet["C1"].value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(1,sheet.max_column+1):
            # dict["name"="heidi"]
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(dict)


