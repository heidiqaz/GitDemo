import openpyxl
class HomePageData:

    test_homepage_data = [{"name": "heidi", "email": "heidi@aaa.com", "gender": "Female"}, {"name": "faye", "email": "faye@aaa.com", "gender": "Male"}]
    @staticmethod
    def get_test_data(test_case_name):
        book = openpyxl.load_workbook("/Users/heidi/PycharmProjects/pythonSelFramework/pyDemo.xlsx")
        sheet = book.active
        dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[dict]
